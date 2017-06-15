import os
import shutil

from openpyxl import load_workbook
from openpyxl.worksheet import Worksheet


def str_to_dict(s):
    if "=" in s:
        l = s.split("=")
    return {l[0]: l[1]}


class ExcelReader(object):
    def __init__(self, excel_path, from_folder, picked_folder, suffix, sheet_name, **kwargs):
        self._excel_path = excel_path
        self._from_folder = from_folder
        self._picked_folder = picked_folder
        self._suffix = suffix
        self._condition_dict = {}
        self._path_col_list = []
        for k in kwargs:
            if "condition" in str(k):
                self._condition_dict.update(str_to_dict(kwargs[k]))
            if "path" in str(k):
                self._path_col_list.append(kwargs[k])
        self._sheet_name = sheet_name

    def read(self):
        excel = load_workbook(self._excel_path)
        sheet = None
        if excel[self._sheet_name]:
            self.__read_sheet(excel[self._sheet_name])

    def __read_sheet(self, sheet: Worksheet):
        path_column_index_dict = {}
        condition_value_index_dict = {}
        head_row_idx = -1
        idx_all_found = False
        for row in sheet.iter_rows():
            picked_path = self._picked_folder
            from_path = self._from_folder
            if not idx_all_found:
                for cell in row:
                    if cell.value in self._path_col_list:
                        path_column_index_dict[cell.value] = cell.column
                    if cell.value in self._condition_dict:
                        condition_value_index_dict[cell.value] = cell.column
                    if len(path_column_index_dict) == len(self._path_col_list) and \
                                    len(condition_value_index_dict) == len(self._condition_dict):
                        head_row_idx = cell.row
                        idx_all_found = True
                        break
            crt_row_idx = row[0].row
            if head_row_idx != -1 and row[0].row > head_row_idx:
                is_target = False
                for condition_column in condition_value_index_dict:
                    sheet_value = sheet[condition_value_index_dict[condition_column] + str(crt_row_idx)].value
                    condition_value = self._condition_dict[condition_column]
                    if sheet_value == condition_value:
                        is_target = True
                    else:
                        is_target = False
                        break
                if is_target:
                    for path_column in path_column_index_dict:
                        s = sheet[path_column_index_dict[path_column] + str(crt_row_idx)].value
                        picked_path = os.path.join(picked_path, s)
                        from_path = os.path.join(from_path, s)
                        if os.path.exists(from_path) and not os.path.exists(picked_path):
                            if not str(s).endswith(self._suffix):
                                os.makedirs(picked_path)
                            else:
                                picked_file_folder = picked_path[:picked_path.rindex("\\")]
                                if not os.path.exists(picked_file_folder):
                                    os.makedirs(picked_file_folder)
                                # shutil.copytree(from_path, picked_path)
                                shutil.copy(from_path, picked_path)
                                print("copied file: " + from_path)
                        if str(s).endswith(self._suffix):
                            if not os.path.exists(picked_path):
                                print("warn : copying file 「" + picked_path + "」 failed!")



if __name__ == "__main__":
    reader = ExcelReader("C:\\Users\\xin.cheng\\Desktop\\temp\\【インテック／タダノ】フォーム変換テスト報告一覧.xlsx", "D:\\16.インテック／タダノ／フォーム変換\\10.xml(修正前)", "C:\\Users\\xin.cheng\\Desktop\\temp\\picked_form", ".xml", "Sheet1", condition1="イメージの種類=実行時に設定",
                         path1="フォルダー", path2="ファイル名")
    reader.read()
